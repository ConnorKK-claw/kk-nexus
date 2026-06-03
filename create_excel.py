import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "14家上海国资案例数据"

# ── Headers ──
headers = ["序号", "公司名称", "草案公告日", "公告后次日股票涨跌幅", "公告后20个交易日股票涨跌幅",
           "第一次解禁时股票涨跌幅", "第一次解禁是否达成", "第二次解禁时股票涨跌幅",
           "第二次解禁是否达成", "第三次解禁时股票涨跌幅", "第三次解禁是否达成"]

# ── Data in reference order ──
data = [
    [1, "上海建科",  "2025-12-16", "+0.2%",  "+4.2%",  "—", "未到期", "—", "未到期", "—", "未到期"],
    [2, "上海机场",  "2024-05-14", "-1.5%",  "-8.4%",  "—", "未到期", "—", "未到期", "—", "未到期"],
    [3, "锦江酒店",  "2024-08-10", "-1.5%",  "-0.6%",  "—", "未到期", "—", "未到期", "—", "未到期"],
    [4, "华建集团②", "2022-01-29", "+1.9%",  "+6.5%",  "-11.7%", "Y", "—", "未到期", "—", "未到期"],
    [5, "外服控股",  "2022-01-28", "+0.5%",  "+1.5%",  "-16.6%", "Y", "-13.0%", "Y", "—", "未到期"],
    [6, "赢合科技②", "2022-11-11", "-1.9%",  "-3.0%",  "-17.9%", "Y", "+37.6%", "Y", "—", "未到期"],
    [7, "东方创业",  "2021-11-30", "+0.6%",  "+5.4%",  "—", "Y", "—", "X", "—", "X"],
    [8, "上港集团",  "2021-04-24", "-1.4%",  "+9.4%",  "+65.6%", "Y", "+52.3%", "Y", "—", "未到期"],
    [9, "申能股份",  "2021-01-26", "+0.8%",  "+4.7%",  "—", "X", "+110.5%", "Y", "+122.6%", "Y"],
    [10, "国泰君安", "2020-06-08", "+1.5%",  "+41.4%", "+2.5%",  "Y", "+5.8%",  "Y", "+30.4%", "Y"],
    [11, "华谊集团", "2020-11-25", "-3.2%",  "-9.6%",  "+3.2%",  "Y", "+23.3%", "Y", "+114.7%","Y"],
    [12, "上海电气", "2019-01-23", "+0.2%",  "+13.3%", "—", "X(终止)", "—", "—", "—", "—"],
    [13, "华建集团①", "2018-12-25", "+0.2%",  "-2.0%",  "-25.3%", "Y", "-3.1%",  "Y", "-21.7%", "Y"],
    [14, "赢合科技①", "2017-10-24", "-0.3%",  "-5.1%",  "-14.9%", "Y", "+3.1%",  "Y", "—", "X"],
]

# ── Styles ──
header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="微软雅黑", size=10)
data_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

# ── Title row ──
ws.merge_cells("A1:K1")
title_cell = ws["A1"]
title_cell.value = "上海国资上市公司股权激励案例 — 股价表现一览"
title_cell.font = Font(name="微软雅黑", bold=True, size=14, color="1F3864")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# ── Subtitle row ──
ws.merge_cells("A2:K2")
sub_cell = ws["A2"]
sub_cell.value = "数据范围：2017-2026年｜14家上海国资上市公司｜共16期股权激励计划"
sub_cell.font = Font(name="微软雅黑", size=9, color="808080", italic=True)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

# ── Write header row (row 3) ──
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws.row_dimensions[3].height = 32

# ── Write data (starting row 4) ──
for row_idx, row_data in enumerate(data, 4):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border
        if (row_idx - 4) % 2 == 1:
            cell.fill = alt_fill
    ws.row_dimensions[row_idx].height = 24

# ── Column widths ──
col_widths = [6, 14, 14, 22, 26, 22, 18, 22, 18, 22, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Freeze panes ──
ws.freeze_panes = "A4"

# ── Auto filter ──
ws.auto_filter.ref = f"A3:K{3+len(data)}"

# ── Note row ──
note_row = 3 + len(data) + 2
ws.merge_cells(f"A{note_row}:K{note_row}")
note = ws.cell(row=note_row, column=1)
note.value = "说明：① 涨跌幅为草案公告后次日/20个交易日的股价变动；② 解禁时涨跌幅为授予价至解禁日收盘价的涨幅；③ ②/① 分别表示该公司的第一期/第二期激励计划；④ X=未达成业绩条件，X(终止)=计划终止"
note.font = Font(name="微软雅黑", size=8, color="808080")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[note_row].height = 30

# ── Legend row ──
legend_row = note_row + 1
ws.merge_cells(f"A{legend_row}:K{legend_row}")
legend = ws.cell(row=legend_row, column=1)
legend.value = "数据来源：Equity-Incentive Vault / 公司公告 / Tushare Pro 行情数据"
legend.font = Font(name="微软雅黑", size=8, color="808080")
legend.alignment = Alignment(horizontal="left", vertical="center")

out_path = "C:\\Users\\hexk\\OneDrive\\文档\\New project 6\\14家上海国资案例数据.xlsx"
wb.save(out_path)
print(f"File saved to {out_path}")
print("Done!")
